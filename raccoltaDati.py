from datetime import datetime, timezone
import openpyxl
import os

def valori():
    with open("values.txt", "r") as file:
        lines = file.readlines()
        value1 = lines[0].strip().split(":")[1].strip()
        value2 = lines[1].strip().split(":")[1].strip()
        value3 = lines[2].strip().split(":")[1].strip()
        value4 = lines[3].strip().split(":")[1].strip()
        return value1, value2, value3,  value4 # cognome, cicli,  sxodx,quadranti


def datiGabor(N,dizionario):
    data = datetime.today().strftime('%d-%m-%Y')
    filename = valori()[0] + str(data)+'.xlsx'
    orario = str(datetime.today().strftime('%H-%M'))
    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
        # Rimuovi il foglio predefinito "sheet"
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

    wb.create_sheet(orario)
    sh1 = wb[orario]
    if valori()[2] == 'left':
        sh1["A1"].value = "Sinistra"
    else:
        sh1["A1"].value = "Destra"
    start_char = "A"
    num_chars = N*2

    for i in range(num_chars):
        casella = start_char + str(i+2)
        sh1[casella].value = 'Quadrante N ' + str((i+1))
    sh1['A' + str(N * 2 + 2)].value = 'Falso Positivo Su'
    sh1['A' + str(N * 2 + 3)].value = 'Falso Positivo Giù'
    sh1['B1'].value = "Giusti"
    sh1['C1'].value = "Sbagliati"
    for i in range(N*2):
        sh1['B'+ str(i+2)].value = dizionario['Q'+str(i+1)]['Giusti']
        sh1['C' + str(i + 2)].value = dizionario['Q' + str(i + 1)]['Sbagliati']
    sh1['B'+  str(N * 2 + 2)].value = dizionario['FalsoPostivoUp']['Giusti']
    sh1['C' + str(N * 2 + 2)].value = dizionario['FalsoPostivoUp']['Sbagliati']
    sh1['B' + str(N * 2 + 3)].value = dizionario['FalsoPostivoDown']['Giusti']
    sh1['C' + str(N * 2 + 3)].value = dizionario['FalsoPostivoDown']['Sbagliati']

    try:
        sh1.column_dimensions['A'].auto_size = True
        sh1.column_dimensions['A'].width += 3
        wb.save(filename)
    except PermissionError:
        print('Chiudere il file excel e ripetere esperimento')



def creazioneDizionario(n):
    n=n*2
    dizionario = {}
    for i in range(n):
        dizionario['Q'+str(i+1)] = {'Giusti':0,'Sbagliati':0}
    dizionario['FalsoPostivoUp'] = {'Giusti':0,'Sbagliati':0}
    dizionario['FalsoPostivoDown'] = {'Giusti': 0, 'Sbagliati': 0}
    return dizionario



